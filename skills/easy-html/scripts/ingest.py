#!/usr/bin/env python3
"""
ingest.py — easy-html 统一输入解析器

把多种输入格式归一化为「结构化文本块」(JSON)，供 agent 据此生成语义化 HTML 骨架。

支持格式：
  .md / .markdown / .txt   → 原文返回（agent 自行理解 markdown 结构）
  .xlsx                    → openpyxl 读取，每个 sheet 转为表格块
  .docx                    → 零依赖解析（unzip + word/document.xml），抓段落 + 表格
  .xls / .doc              → 降级：不解析，返回提示让用户另存为 xlsx/docx 或粘贴文本

输出 JSON 结构：
{
  "ok": true,
  "source": "<文件路径>",
  "kind": "markdown|text|xlsx|docx",
  "title_hint": "<从文件名/首行/首标题猜的标题>",
  "blocks": [
    {"type": "heading", "level": 1, "text": "..."},
    {"type": "paragraph", "text": "..."},
    {"type": "table", "caption": "Sheet1", "header": [...], "rows": [[...], ...]},
    {"type": "raw", "text": "<原始 markdown/纯文本全文>"}
  ],
  "notes": ["..."]      # 解析过程中的提示/降级说明
}

用法：
  python3 ingest.py <file>                 # 解析文件
  python3 ingest.py --text "一段文字"        # 直接解析纯文本
  python3 ingest.py --stdin                 # 从 stdin 读纯文本

退出码：0 成功；1 解析失败（含降级提示，ok=false）
"""
import argparse
import json
import os
import re
import sys
import zipfile
import html as _html
from xml.etree import ElementTree as ET


def _result(ok, source, kind, blocks, title_hint="", notes=None):
    return {
        "ok": ok,
        "source": source,
        "kind": kind,
        "title_hint": title_hint,
        "blocks": blocks or [],
        "notes": notes or [],
    }


def _title_from_filename(path):
    base = os.path.basename(path)
    name = os.path.splitext(base)[0]
    return name.strip()


# ── markdown / 纯文本 ──────────────────────────────────────────────
def parse_text(text, source="<text>", kind="text"):
    """纯文本/markdown：返回 raw 全文 + 猜标题。

    不做深度结构化解析 —— markdown 的语义结构由 agent 直接理解更灵活可靠。
    只额外抽一个 title_hint 方便上层设置 <title>。
    """
    title = ""
    # markdown 一级标题优先
    m = re.search(r"^\s*#\s+(.+)$", text, re.MULTILINE)
    if m:
        title = m.group(1).strip()
    else:
        # 否则取第一行非空文本（截断）
        for line in text.splitlines():
            s = line.strip().lstrip("#").strip()
            if s:
                title = s[:40]
                break
    blocks = [{"type": "raw", "text": text}]
    return _result(True, source, kind, blocks, title_hint=title)


# ── xlsx（openpyxl）─────────────────────────────────────────────────
def parse_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _result(
            False, path, "xlsx", [],
            notes=["当前环境缺少 openpyxl，无法解析 .xlsx。"
                   "请让用户把数据粘贴为文本/markdown 表格，或安装 openpyxl。"],
        )
    notes = []
    blocks = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return _result(False, path, "xlsx", [],
                       notes=[f"打开 xlsx 失败：{e}"])
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=True):
            # 跳过整行全空
            if row is None or all(c is None or str(c).strip() == "" for c in row):
                continue
            rows.append(["" if c is None else str(c) for c in row])
        if not rows:
            continue
        # 第一行当表头
        header = rows[0]
        body = rows[1:]
        # 截断超大表（提示）
        MAX_ROWS = 500
        if len(body) > MAX_ROWS:
            notes.append(f"工作表「{ws.title}」共 {len(body)} 行，已截断展示前 {MAX_ROWS} 行。")
            body = body[:MAX_ROWS]
        blocks.append({
            "type": "table",
            "caption": ws.title,
            "header": header,
            "rows": body,
        })
    wb.close()
    if not blocks:
        return _result(False, path, "xlsx", [],
                       notes=["xlsx 中没有读到任何非空数据。"])
    return _result(True, path, "xlsx", blocks,
                   title_hint=_title_from_filename(path), notes=notes)


# ── docx（零依赖：unzip + 解析 word/document.xml）────────────────────
_W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"


def _w(tag):
    return f"{{{_W_NS}}}{tag}"


def _para_text(p):
    """提取一个 <w:p> 段落的纯文本。"""
    parts = []
    for t in p.iter(_w("t")):
        parts.append(t.text or "")
    return "".join(parts).strip()


def _para_style_level(p):
    """从段落样式名猜 heading level，返回 0 表示普通段落。"""
    ppr = p.find(_w("pPr"))
    if ppr is None:
        return 0
    pstyle = ppr.find(_w("pStyle"))
    if pstyle is None:
        return 0
    val = pstyle.get(_w("val")) or ""
    # Word 标题样式常见：Heading1 / Heading2 / 标题 1 ...
    m = re.search(r"(?:Heading|标题)\s*([1-6])", val)
    if m:
        return int(m.group(1))
    if val.lower() in ("title",):
        return 1
    return 0


def parse_docx(path):
    notes = []
    try:
        zf = zipfile.ZipFile(path)
    except Exception as e:  # noqa: BLE001
        return _result(False, path, "docx", [],
                       notes=[f"打开 docx 失败（可能不是有效 .docx）：{e}"])
    try:
        with zf:
            if "word/document.xml" not in zf.namelist():
                return _result(False, path, "docx", [],
                               notes=["docx 内未找到 word/document.xml，可能是损坏文件。"])
            xml = zf.read("word/document.xml")
    except Exception as e:  # noqa: BLE001
        return _result(False, path, "docx", [], notes=[f"读取 document.xml 失败：{e}"])

    try:
        root = ET.fromstring(xml)
    except ET.ParseError as e:
        return _result(False, path, "docx", [], notes=[f"解析 document.xml XML 失败：{e}"])

    body = root.find(_w("body"))
    if body is None:
        return _result(False, path, "docx", [], notes=["docx body 为空。"])

    blocks = []
    title_hint = ""
    # 按文档顺序遍历 body 直接子节点：段落(w:p) 和 表格(w:tbl)
    for el in list(body):
        tag = el.tag
        if tag == _w("p"):
            txt = _para_text(el)
            if not txt:
                continue
            lvl = _para_style_level(el)
            if lvl > 0:
                blocks.append({"type": "heading", "level": lvl, "text": txt})
                if not title_hint and lvl == 1:
                    title_hint = txt
            else:
                blocks.append({"type": "paragraph", "text": txt})
        elif tag == _w("tbl"):
            tbl_rows = []
            for tr in el.findall(_w("tr")):
                cells = []
                for tc in tr.findall(_w("tc")):
                    # 单元格可能含多段落，拼接
                    cell_txt = " ".join(
                        _para_text(p) for p in tc.findall(_w("p"))
                    ).strip()
                    cells.append(cell_txt)
                if cells:
                    tbl_rows.append(cells)
            if tbl_rows:
                header = tbl_rows[0]
                rows = tbl_rows[1:]
                blocks.append({
                    "type": "table",
                    "caption": "",
                    "header": header,
                    "rows": rows,
                })

    if not blocks:
        return _result(False, path, "docx", [],
                       notes=["docx 中未提取到任何文本或表格内容。"])
    if not title_hint:
        # 退化为第一个 heading 或文件名
        for b in blocks:
            if b["type"] == "heading":
                title_hint = b["text"]
                break
        if not title_hint:
            title_hint = _title_from_filename(path)
    return _result(True, path, "docx", blocks, title_hint=title_hint, notes=notes)


# ── 老格式降级 ──────────────────────────────────────────────────────
def degrade_legacy(path, ext):
    fmt = "xls" if ext == ".xls" else "doc"
    target = "xlsx" if fmt == "xls" else "docx"
    return _result(
        False, path, fmt, [],
        notes=[
            f"检测到老版 .{fmt} 二进制格式，easy-html 不直接解析（避免引入重依赖）。",
            f"请用户把文件「另存为」.{target} 后重新提供；或直接把内容粘贴为文本/markdown。",
        ],
    )


# ── 入口 ────────────────────────────────────────────────────────────
def ingest_file(path):
    if not os.path.exists(path):
        return _result(False, path, "unknown", [], notes=[f"文件不存在：{path}"])
    ext = os.path.splitext(path)[1].lower()
    if ext in (".md", ".markdown"):
        with open(path, encoding="utf-8", errors="replace") as f:
            return parse_text(f.read(), source=path, kind="markdown")
    if ext == ".txt":
        with open(path, encoding="utf-8", errors="replace") as f:
            return parse_text(f.read(), source=path, kind="text")
    if ext == ".xlsx":
        return parse_xlsx(path)
    if ext == ".docx":
        return parse_docx(path)
    if ext in (".xls", ".doc"):
        return degrade_legacy(path, ext)
    # 未知扩展名：当纯文本试读
    try:
        with open(path, encoding="utf-8", errors="replace") as f:
            return parse_text(f.read(), source=path, kind="text")
    except Exception as e:  # noqa: BLE001
        return _result(False, path, "unknown", [],
                       notes=[f"无法识别的文件类型 {ext}，按文本读取也失败：{e}"])


def main():
    ap = argparse.ArgumentParser(description="easy-html 统一输入解析器")
    ap.add_argument("file", nargs="?", help="要解析的文件路径")
    ap.add_argument("--text", help="直接解析一段纯文本")
    ap.add_argument("--stdin", action="store_true", help="从 stdin 读纯文本")
    args = ap.parse_args()

    if args.text is not None:
        res = parse_text(args.text, source="<text>", kind="text")
    elif args.stdin:
        res = parse_text(sys.stdin.read(), source="<stdin>", kind="text")
    elif args.file:
        res = ingest_file(args.file)
    else:
        ap.error("需要提供 <file> 或 --text 或 --stdin")
        return

    print(json.dumps(res, ensure_ascii=False, indent=2))
    sys.exit(0 if res["ok"] else 1)


if __name__ == "__main__":
    main()
